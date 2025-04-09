import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import argparse
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.ensemble import RandomForestRegressor, AdaBoostRegressor, StackingRegressor, BaggingRegressor
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split, GridSearchCV, KFold
from xgboost import XGBRegressor
from sklearn.metrics import make_scorer
from sklearn.metrics import mean_absolute_percentage_error as mape
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import NearestNeighbors
from sklearn.model_selection import cross_val_score
from sklearn.tree import DecisionTreeRegressor
from sklearn.svm import SVR
from sklearn.impute import KNNImputer
from openpyxl import load_workbook
import itertools

def prepare_df(path):
    df = pd.read_excel(path)
    df.columns = df.iloc[0]
    df = df[1:]
    
    # Map age group to integer
    avg_age_mapping = {
        '0-14': 7,
        '15-24': 20,
        '25-44': 33,
        '45-64': 55,
        '65+': 75
    }
    
    # Map gender to integer
    gender_mapping = {
        'Male': 0,
        'Female': 1
    }

    # Map ethnicity to integer
    ethnicity_mapping = {
        'Māori': 0,
        'non-Māori': 1
    }
    
    # Apply the mapping to the 'Age_Group' column
    df['average_age'] = df['age'].map(avg_age_mapping)
    df['gender_idx'] = df['gender'].map(gender_mapping)
    df['ethnicity_idx'] = df['ethnicity'].map(ethnicity_mapping)
    
    # Impute missing values in 'average_age' with the mean
    df['average_age'] = df['average_age'].fillna(df['average_age'].mean())
    
    # Impute missing values in 'gender_idx' and 'ethnicity_idx' with the mode
    df['gender_idx'] = df['gender_idx'].fillna(df['gender_idx'].mode()[0])
    df['ethnicity_idx'] = df['ethnicity_idx'].fillna(df['ethnicity_idx'].mode()[0])
    
    # Convert the specified columns to floats
    df[['tax_increase', 'outlet_reduction', 'dec_smoking_prevalence', 
        'dec_tobacco_supply', 'dec_smoking_uptake', 'qalys_pc']] = df[['tax_increase', 'outlet_reduction', 
        'dec_smoking_prevalence', 'dec_tobacco_supply', 'dec_smoking_uptake', 'qalys_pc']].apply(pd.to_numeric, errors='coerce').astype('float')
    
    # Columns to be used for model building
    df_vape = df[['tax_increase', 'outlet_reduction', 'dec_smoking_prevalence', 
                  'dec_tobacco_supply', 'dec_smoking_uptake', 'average_age', 
                  'gender_idx', 'ethnicity_idx', 'qalys_pc', 'hs_costs_pc']]
    return df_vape

def introduce_std_noise(df, columns):
    """
    Introduces noise into the specified columns by adding a random perturbation 
    based on half the standard deviation of each column.
    
    Parameters:
    df (pd.DataFrame): The input dataframe.
    columns (list): List of column names to introduce noise.
    seed (int, optional): Random seed for reproducibility.
    
    Returns:
    pd.DataFrame: A new dataframe with noise introduced.
    """    
    df_copy = df.copy()
    for col in columns:
        std_dev = df[col].std() / 2  # Compute half the standard deviation
        noise = np.random.normal(0, std_dev, size=df.shape[0])  # Generate noise
        df_copy[col] += noise  # Add noise to the column
    
    return df_copy
def simple_duplicate(X, y, n_samples=200, std_dev=0.4, random_state=None):
    """
    Duplicate X and y with added Gaussian noise.
    
    Parameters:
        X (numpy.ndarray): Feature matrix.
        y (numpy.ndarray): Target values.
        n_samples (int): Number of new samples to generate.
        std_dev (float): Standard deviation of the noise.
        random_state (int, optional): Seed for reproducibility.

    Returns:
        X_aug (numpy.ndarray): Augmented feature matrix.
        y_aug (numpy.ndarray): Augmented target values.
    """
    if random_state is not None:
        np.random.seed(random_state)
    
    # Select random samples from X and y
    indices = np.random.choice(X.shape[0], size=n_samples, replace=True)
    X_selected = X[indices]
    y_selected = y[indices]
    
    # Add Gaussian noise to X
    noise = np.random.normal(loc=0.0, scale=std_dev, size=X_selected.shape)
    X_aug = X_selected + noise
    
    # Keep y unchanged or add noise to y if needed (optional)
    y_aug = y_selected  # Add noise to y if required: y_selected + np.random.normal(0, std_dev, size=y_selected.shape)
    
    return X_aug, y_aug

# Function to generate synthetic samples
def knn_samples(X, y, n_samples, random_state=42):
    np.random.seed(random_state)
    nn = NearestNeighbors(n_neighbors=5)
    nn.fit(X)

    synthetic_X = []
    synthetic_y = []
    for _ in range(n_samples):
        idx = np.random.randint(0, len(X))
        neighbors = nn.kneighbors([X[idx]], return_distance=False)[0]
        
        neighbor_idx = np.random.choice(neighbors)
        lam = np.random.uniform(0, 1)
        
        # Generate synthetic sample using interpolation
        new_sample_X = X[idx] + lam * (X[neighbor_idx] - X[idx])
        new_sample_y = y[idx] + lam * (y[neighbor_idx] - y[idx])
        
        synthetic_X.append(new_sample_X)
        synthetic_y.append(new_sample_y)
    
    return np.array(synthetic_X), np.array(synthetic_y)

def knn_samples_rebalanced(X, y, n_samples, threshold=0.2, random_state=42):
    np.random.seed(random_state)
    nn = NearestNeighbors(n_neighbors=5)
    nn.fit(X)

    # Calculate distances to the nearest neighbors
    distances, _ = nn.kneighbors(X)

    # Calculate the number of neighbors within the threshold for each sample
    neighbors_within_threshold = np.sum(distances[:, 1:] < threshold, axis=1)

    # Calculate inverse probabilities
    selection_probabilities = 1 / (neighbors_within_threshold + 1e-6)  # Add small constant to avoid division by zero
    selection_probabilities /= selection_probabilities.sum()  # Normalize to sum to 1

    synthetic_X = []
    synthetic_y = []

    for _ in range(n_samples):
        # Select a sample based on the weighted probability distribution
        idx = np.random.choice(len(X), p=selection_probabilities)

        # Select a random neighbor of the chosen sample
        neighbors = nn.kneighbors([X[idx]], return_distance=False)[0]
        neighbor_idx = np.random.choice(neighbors)

        # Linear interpolation for synthetic sample generation
        lam = np.random.uniform(0, 1)
        new_sample_X = X[idx] + lam * (X[neighbor_idx] - X[idx])
        new_sample_y = y[idx] + lam * (y[neighbor_idx] - y[idx])

        synthetic_X.append(new_sample_X)
        synthetic_y.append(new_sample_y)

    # Stack the original and synthetic data
    X_full = np.vstack([X_train.values, synthetic_X])
    y_full = np.hstack([y_flat, synthetic_y])
    
    return np.array(X_full), np.array(y_full)
    
def build_lr(X_train, X_test, y_train, y_test):
    # Initialize the Linear Regression model
    model = LinearRegression()
    
    # Fit the model to the training data
    model.fit(X_train, y_train)
    
    # Make predictions on the test set
    y_pred = model.predict(X_test)
    
    # Evaluate the model
    lr_mape = mape(y_test, y_pred)

    print("Test MAPE for Linear Regression Model:", lr_mape)
    
    return model, lr_mape

def build_no_bootstrap(X_train, X_test, y_train, y_test, cv=5):    
    # Define the RandomForestRegressor model with bootstrap disabled
    rf_model_no_bootstrap = RandomForestRegressor(random_state=42, bootstrap=False)
    
    # Define the parameter grid to search over
    param_grid_no_bootstrap = {
        'n_estimators': [100, 200, 300],    # Number of trees in the forest
        'max_depth': [3, 5, 10],            # Maximum depth of the tree
        'min_samples_leaf': [1, 2, 4],      # Minimum number of samples required to be at a leaf node
    }
    
    # Define the MAPE scorer (using Mean Absolute Percentage Error)
    mape_scorer = make_scorer(mape, greater_is_better=False)
    
    # Setup GridSearchCV to perform cross-validation
    grid_search_rf_no_bootstrap = GridSearchCV(estimator=rf_model_no_bootstrap, param_grid=param_grid_no_bootstrap, 
                                               scoring=mape_scorer, cv=5, verbose=1, n_jobs=-1)
    
    # Fit the grid search to the duplicated training data
    grid_search_rf_no_bootstrap.fit(X_train, y_train)
    
    # Best hyperparameters from grid search
    print("Best Parameters for Random Forest (No Bootstrap):", grid_search_rf_no_bootstrap.best_params_)
    
    # Best MAPE score from cross-validation
    print("Best CV MAPE for Random Forest (No Bootstrap):", -grid_search_rf_no_bootstrap.best_score_)
    
    # Train a final model using the best parameters
    best_no_bootstrap_model = grid_search_rf_no_bootstrap.best_estimator_
    
    # Evaluate on the test set
    y_pred_rf_no_bootstrap = best_no_bootstrap_model.predict(X_test)
    
    # Calculate the test MAPE
    test_mape = mape(y_test, y_pred_rf_no_bootstrap)
    print("Test MAPE for Random Forest (No Bootstrap):", test_mape)

    return best_no_bootstrap_model, test_mape

def build_bootstrap(X_train, X_test, y_train, y_test, cv=5):    
    # Define the RandomForestRegressor model
    rf_model = RandomForestRegressor(random_state=42, bootstrap=True)
    
    # Define the parameter grid to search over
    param_grid = {
        'n_estimators': [100, 200, 300],    # Number of trees in the forest
        'max_depth': [5, 10, 20],            # Maximum depth of the tree
        'min_samples_leaf': [1, 5, 10],      # Minimum number of samples required to be at a leaf node
        'max_samples': [0.5, 0.7, 1.0],     # Maximum number of samples to draw from the data with replacement
    }
    
    # Define the MAPE scorer (using Mean Absolute Percentage Error)
    mape_scorer = make_scorer(mape, greater_is_better=False)
    
    # Setup GridSearchCV to perform cross-validation
    grid_search_rf = GridSearchCV(estimator=rf_model, param_grid=param_grid, 
                                  scoring=mape_scorer, cv=5, verbose=1, n_jobs=-1)
    
    # Fit the grid search to the duplicated training data
    grid_search_rf.fit(X_train, y_train)
    
    # Best hyperparameters from grid search
    print("Best Parameters for Random Forest:", grid_search_rf.best_params_)
    
    # Best MAPE score from cross-validation
    print("Best CV MAPE for Random Forest:", -grid_search_rf.best_score_)
    
    # Train a final model using the best parameters
    best_bootstrap_model = grid_search_rf.best_estimator_
    
    # Evaluate on the test set
    y_pred_rf = best_bootstrap_model.predict(X_test)
    
    # Calculate the test MAPE
    test_mape = mape(y_test, y_pred_rf)
    print("Test MAPE for Random Forest:", test_mape)

    return best_bootstrap_model, test_mape

def build_xgboost(X_train, X_test, y_train, y_test, cv=5):
    # Define the XGBoost model
    xgb_model = XGBRegressor(objective='reg:squarederror', random_state=42)
    
    # Define the parameter grid to search over
    param_grid = {
        'n_estimators': [100, 200, 300],   # Number of trees
        'max_depth': [5, 10, 20],            # Depth of the trees
        'min_child_weight': [1, 5, 10],     # Minimum sum of instance weight (hessian)
        'reg_lambda': [0.01, 0.1, 1, 10],  # L2 regularization term (lambda)
        'reg_alpha': [0.01, 0.1, 1, 10],      # L1 regularization term (alpha)
    }
    
    # Define the MAPE scorer (as we are optimizing based on Mean Absolute Percentage Error)
    mape_scorer = make_scorer(mape, greater_is_better=False)
    
    # Setup GridSearchCV to perform cross-validation
    grid_search = GridSearchCV(estimator=xgb_model, param_grid=param_grid, 
                               scoring=mape_scorer, cv=5, verbose=1, n_jobs=-1)
    
    # Fit the grid search to the duplicated training data
    grid_search.fit(X_train, y_train)
    
    # Best hyperparameters from grid search
    print("Best Parameters for XGBoost:", grid_search.best_params_)
    
    # Best MAPE score from cross-validation
    print("Best CV MAPE for XGBoost:", -grid_search.best_score_)
    
    # Train a final model using the best parameters
    best_xgb_model = grid_search.best_estimator_
    
    # Evaluate on the test set
    y_pred = best_xgb_model.predict(X_test)
    
    # Calculate the test MAPE
    test_mape = mape(y_test, y_pred)
    print("Test MAPE for XGBoost:", test_mape)

    return best_xgb_model, test_mape

def build_xgboost_with_bootstrap(X_train, X_test, y_train, y_test, cv=5):
    # Define the XGBoost model
    xgb_model = XGBRegressor(objective='reg:squarederror', random_state=42)
    
    # Define the parameter grid to search over
    param_grid = {
        'n_estimators': [100, 200, 300],   # Number of trees
        'max_depth': [5, 10, 20],            # Depth of the trees
        'min_child_weight': [1, 5, 10],     # Minimum sum of instance weight (hessian)
        'reg_lambda': [0.01, 0.1, 1, 10],  # L2 regularization term (lambda)
        'reg_alpha': [0.01, 0.1, 1, 10],    # L1 regularization term (alpha)
        'subsample': [0.5, 0.75, 0.999],      # Fraction of samples used for each tree (bootstrap)
        'sampling_method': ['uniform']      # Ensures sampling with replacement (bootstrapping)
    }
    
    # Define the MAPE scorer (as we are optimizing based on Mean Absolute Percentage Error)
    mape_scorer = make_scorer(mape, greater_is_better=False)
    
    # Setup GridSearchCV to perform cross-validation
    grid_search = GridSearchCV(estimator=xgb_model, param_grid=param_grid, 
                               scoring=mape_scorer, cv=cv, verbose=1, n_jobs=-1)
    
    # Fit the grid search to the training data
    grid_search.fit(X_train, y_train)
    
    # Best hyperparameters from grid search
    print("Best Parameters for XGBoost:", grid_search.best_params_)
    
    # Best MAPE score from cross-validation
    print("Best CV MAPE for XGBoost:", -grid_search.best_score_)
    
    # Train a final model using the best parameters
    best_xgb_model = grid_search.best_estimator_
    
    # Evaluate on the test set
    y_pred = best_xgb_model.predict(X_test)
    
    # Calculate the test MAPE
    test_mape = mape(y_test, y_pred)
    print("Test MAPE for XGBoost:", test_mape)

    return best_xgb_model, test_mape

def build_adaboost(X_train, X_test, y_train, y_test, cv=5):
    # Define the AdaBoost model
    ada_model = AdaBoostRegressor(random_state=42)
    
    # Define the parameter grid to search over
    param_grid = {
        'n_estimators': [50, 100, 200],  # Number of boosting stages
        'learning_rate': [0.01, 0.1, 1],  # Weight applied to each regressor at each iteration
        'loss': ['linear', 'square', 'exponential']  # Loss function to optimize
    }
    
    # Define the MAPE scorer
    mape_scorer = make_scorer(mape, greater_is_better=False)
    
    # Setup GridSearchCV to perform cross-validation
    grid_search = GridSearchCV(estimator=ada_model, param_grid=param_grid, 
                               scoring=mape_scorer, cv=cv, verbose=1, n_jobs=-1)
    
    # Fit the grid search to the training data
    grid_search.fit(X_train, y_train)
    
    # Best hyperparameters from grid search
    print("Best Parameters for AdaBoost:", grid_search.best_params_)
    
    # Best MAPE score from cross-validation
    print("Best CV MAPE for AdaBoost:", -grid_search.best_score_)
    
    # Train a final model using the best parameters
    best_ada_model = grid_search.best_estimator_
    
    # Evaluate on the test set
    y_pred = best_ada_model.predict(X_test)
    
    # Calculate the test MAPE
    test_mape = mape(y_test, y_pred)
    print("Test MAPE for AdaBoost:", test_mape)
    
    return best_ada_model, test_mape

def build_adaboost_with_bootstrap(X_train, X_test, y_train, y_test, cv=5):
    # Base AdaBoost model
    base_ada = AdaBoostRegressor(random_state=42)
    
    # Wrap AdaBoost in BaggingRegressor for bootstrapping
    bagged_ada = BaggingRegressor(
        estimator=base_ada,
        bootstrap=True,          # Enable sampling with replacement
        random_state=42,
        n_jobs=-1
    )
    
    # Define the parameter grid
    param_grid = {
        'estimator__n_estimators': [50, 100, 200],       # AdaBoost hyperparameters
        'estimator__learning_rate': [0.01, 0.1, 1],
        'estimator__loss': ['linear', 'square', 'exponential'],
        'max_samples': [1.0]                  # Bootstrap sample size
    }
    
    # MAPE scorer
    mape_scorer = make_scorer(mape, greater_is_better=False)
    
    # GridSearchCV
    grid_search = GridSearchCV(
        estimator=bagged_ada,
        param_grid=param_grid,
        scoring=mape_scorer,
        cv=cv,
        verbose=1,
        n_jobs=-1
    )
    
    # Fit the model
    grid_search.fit(X_train, y_train)
    
    # Best hyperparameters
    print("Best Parameters:", grid_search.best_params_)
    print("Best CV MAPE:", -grid_search.best_score_)
    
    # Final model
    best_model = grid_search.best_estimator_
    y_pred = best_model.predict(X_test)
    test_mape = mape(y_test, y_pred)
    print("Test MAPE:", test_mape)
    
    return best_model, test_mape

def build_stacking(X_train, X_test, y_train, y_test, cv=5, tol=1e-4, max_iter=50, patience=5):
    # Define base learners
    base_learners = [
        ('ridge', Ridge(alpha=1.0)),
        ('lasso', Lasso(alpha=0.1)),
        ('svr', SVR(kernel='linear', C=1.0)),
        ('tree', DecisionTreeRegressor(max_depth=5)),
        ('rf', RandomForestRegressor(n_estimators=100, random_state=42))
    ]
    
    # Define meta-learners to try
    meta_learners = {
        'ridge': Ridge(alpha=1.0),
        'lasso': Lasso(alpha=0.1),
        # 'svr': SVR(kernel='linear', C=1.0),
        'rf': RandomForestRegressor(n_estimators=50, random_state=42)
    }
    
    best_mape = float('inf')
    best_model = None
    best_meta = None
    
    for meta_name, meta_learner in meta_learners.items():
        print(f"\nTrying meta-learner: {meta_name}")
        stacking_model = StackingRegressor(estimators=base_learners, final_estimator=meta_learner, cv=cv)
        
        prev_loss = np.inf
        no_improve_count = 0  # Track consecutive iterations with no improvement
        
        for i in range(max_iter):
            stacking_model.fit(X_train, y_train)
            y_pred = stacking_model.predict(X_test)
            test_mape = mape(y_test, y_pred)
            
            print(f"Iteration {i+1}: Test MAPE = {test_mape:.6f}")
            
            if abs(prev_loss - test_mape) < tol:
                print(f"Converged after {i+1} iterations for meta-learner: {meta_name}")
                break
            
            if test_mape >= prev_loss:  
                no_improve_count += 1
            else:
                no_improve_count = 0  # Reset counter if improvement occurs
            
            # If no improvement after 'patience' iterations, stop early
            if no_improve_count >= patience:
                print(f"Early stopping: No improvement after {patience} iterations for {meta_name}")
                break
            
            prev_loss = test_mape
        
        print(f"Final Test MAPE for {meta_name}: {test_mape:.6f}")
        
        if test_mape < best_mape:
            best_mape = test_mape
            best_model = stacking_model
            best_meta = meta_name
    
    print(f"\nBest meta-learner: {best_meta} with Test MAPE: {best_mape:.6f}")
    
    return best_model, best_mape


def build_stacking_with_bootstrap(X_train, X_test, y_train, y_test, cv=5, tol=1e-4, max_iter=50, patience=5, bootstrap_size=1.0):
    # Define base learners
    base_learners = [
        ('ridge', Ridge(alpha=1.0)),
        ('lasso', Lasso(alpha=0.1)),
        ('svr', SVR(kernel='linear', C=1.0)),
        ('tree', DecisionTreeRegressor(max_depth=5)),
        ('rf', RandomForestRegressor(n_estimators=100, random_state=42))
    ]
    
    # Define meta-learners to try
    meta_learners = {
        'ridge': Ridge(alpha=1.0),
        'lasso': Lasso(alpha=0.1),
        # 'svr': SVR(kernel='linear', C=1.0),
        'rf': RandomForestRegressor(n_estimators=50, random_state=42)
    }
    
    best_mape = float('inf')
    best_model = None
    best_meta = None
    
    for meta_name, meta_learner in meta_learners.items():
        print(f"\nTrying meta-learner: {meta_name}")
        stacking_model = StackingRegressor(estimators=base_learners, final_estimator=meta_learner, cv=cv)
        
        prev_loss = np.inf
        no_improve_count = 0
        
        for i in range(max_iter):
            # Bootstrapping the training data
            n_samples = int(len(X_train) * bootstrap_size)
            indices = np.random.choice(len(X_train), size=n_samples, replace=True)
            X_bootstrap = X_train[indices]
            y_bootstrap = y_train[indices]
            
            # Train on the bootstrapped sample
            stacking_model.fit(X_bootstrap, y_bootstrap)
            y_pred = stacking_model.predict(X_test)
            test_mape = mape(y_test, y_pred)
            
            print(f"Iteration {i+1}: Test MAPE = {test_mape:.6f}")
            
            if abs(prev_loss - test_mape) < tol:
                print(f"Converged after {i+1} iterations for meta-learner: {meta_name}")
                break
            
            if test_mape >= prev_loss:
                no_improve_count += 1
            else:
                no_improve_count = 0
            
            if no_improve_count >= patience:
                print(f"Early stopping: No improvement after {patience} iterations for {meta_name}")
                break
            
            prev_loss = test_mape
        
        print(f"Final Test MAPE for {meta_name}: {test_mape:.6f}")
        
        if test_mape < best_mape:
            best_mape = test_mape
            best_model = stacking_model
            best_meta = meta_name
    
    print(f"\nBest meta-learner: {best_meta} with Test MAPE: {best_mape:.6f}")
    
    return best_model, best_mape

def ensemble(models, X_test, y_test):
    predictions = []  # Use list instead of np.array
    for model in models:
        pred = model.predict(X_test)  # Predict on X_test
        predictions.append(pred)  # Append predictions

    predictions = np.array(predictions)  # Convert list to numpy array
    final_pred = np.mean(predictions, axis=0)  # Average predictions across models
    test_mape = mape(y_test, final_pred)  # Compute MAPE

    return test_mape
    
def prepare_df_vape():
    df_vape = pd.read_excel('./Datasets/vaping_data.xlsx')

    # Map age group to integer
    avg_age_mapping = {
        '0-14': 7,
        '15-24': 20,
        '25-44': 33,
        '45-64': 55,
        '65+': 75
    }
    
    # Map gender to integer
    gender_mapping = {
        'Male': 0,
        'Female': 1
    }

    # Map ethnicity to integer
    ethnicity_mapping = {
        'Māori': 0,
        'non-Māori': 1
    }
    
    # Apply the mapping to the 'Age_Group' column
    df_vape['average_age'] = df_vape['age'].map(avg_age_mapping)
    df_vape['gender_idx'] = df_vape['gender'].map(gender_mapping)
    df_vape['ethnicity_idx'] = df_vape['ethnicity'].map(ethnicity_mapping)

    return df_vape

def append_to_excel(file_path, new_row, sheet_name="Sheet1"):
    try:
        # Load the existing workbook
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay")
        
        # Get the last row number
        sheet = book[sheet_name]
        last_row = sheet.max_row

        # Convert new_row to DataFrame and write at the next available row
        df_new = pd.DataFrame([new_row])
        df_new.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=last_row)

        writer.close()
    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        df_new = pd.DataFrame([new_row])
        df_new.to_excel(file_path, index=False, sheet_name=sheet_name)

def impute_mean(df):
    # Replace 'NA' strings with actual NaN values
    df.replace('NA', np.nan, inplace=True)

    # Convert columns to numeric (if they were stored as strings)
    df = df.apply(pd.to_numeric)

    # Impute missing values with column means
    df.fillna(df.mean(numeric_only=True), inplace=True)
    return df

def impute_knn(df, n_neighbors=5):
    # Replace 'NA' strings with actual NaN values
    df.replace('NA', np.nan, inplace=True)

    # Convert columns to numeric (if they were stored as strings)
    df = df.apply(pd.to_numeric)

    # Apply KNN imputation
    imputer = KNNImputer(n_neighbors=n_neighbors)
    df[:] = imputer.fit_transform(df)

    return df

if __name__ == "__main__":
    # Parse Arguments
    parser = argparse.ArgumentParser(description="Select Model to Build")
    
    # Positional arguments for team names
    parser.add_argument('model', type=int, help='0: QALY with base parameters, 1: HSCs with base parameters')
    # Optional flag for variance
    parser.add_argument('--variance', action='store_true', help='Apply variance to the specified columns')
    parser.add_argument('--noimpute', action='store_true', help='Apply variance to the specified columns')

    args = parser.parse_args()        

    # Prepare DataFrame
    print("Preparing DataFrame")
    path = './Datasets/tobacco_data_v2.xlsx'
    if args.noimpute:
        path = './Datasets/tobacco_data.xlsx'
        
    df = prepare_df(path)
    
    columns = ['tax_increase', 'outlet_reduction', 'dec_smoking_prevalence', 
              'dec_tobacco_supply', 'dec_smoking_uptake', 'average_age', 
              'gender_idx', 'ethnicity_idx']

    if args.noimpute:
        df = impute_knn(df)
        # df = impute_mean(df)
   
    if args.model == 0:
        X = df[columns]
        y = df[['qalys_pc']]
    elif args.model == 1:
        X = df[columns]
        y = df[['hs_costs_pc']]
    elif args.model == 2:
        columns.append('qalys_pc')
        X = df[columns]
        y = df[['hs_costs_pc']]
    else:
        raise ValueError("Invalid value for 'args.model'. Expected 0, 1, or 2.")


    # Train Test Split
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)
    y_flat = y_train.values.flatten() # Ensure that y is a 1D array for compatibility

     # Change 
    columns_to_modify = ['dec_smoking_prevalence', 'dec_smoking_uptake']
    if args.variance:
        X_train = introduce_std_noise(X_train, columns_to_modify)

    # Generate synthetic samples
    print("Generating Synthetic Samples")
    X_sim, y_sim = simple_duplicate(X_train.values, y_flat, n_samples=300, random_state=42)
    X_sim = pd.DataFrame(X_sim, columns=columns)
    X_knn, y_knn = knn_samples(X_train.values, y_flat, n_samples=300, random_state=42)
    X_knn = pd.DataFrame(X_knn, columns=columns)
    X_reb, y_reb = knn_samples_rebalanced(X_train.values, y_flat, n_samples=300, random_state=42)
    X_reb = pd.DataFrame(X_reb, columns=columns)

    # Build Models
    # print("Building Stacking Model with no Upsampling")
    # stacking_model, stacking_test_mape = build_stacking(X_train, X_test, y_flat, y_test)
    # print("")

    # print("Building Stacking Model with Simple Duplication")
    # stacking_sim_model, stacking_sim_test_mape = build_stacking(X_sim, X_test, y_sim, y_test)
    # print("")

    # print("Building Stacking Model with KNN Upsampling")
    # stacking_knn_model, stacking_knn_test_mape = build_stacking(X_knn, X_test, y_knn, y_test)
    # print("")

    # print("Building Stacking Model with Rebalanced KNN Upsampling")
    # stacking_reb_model, stacking_reb_test_mape = build_stacking(X_reb, X_test, y_reb, y_test)
    # print("")
    
    # print("Building Linear Regression Model with no Upsampling")
    # lr_model, lr_test_mape = build_lr(X_train, X_test, y_flat, y_test)
    # print("")
    
    # print("Building Linear Regression Model with Simple Duplication")
    # lr_sim_model, lr_sim_test_mape = build_lr(X_sim, X_test, y_sim, y_test)
    # print("")

    # print("Building Linear Regression Model with KNN Upsampling")
    # lr_knn_model, lr_knn_test_mape = build_lr(X_knn, X_test, y_knn, y_test)
    # print("")

    # print("Building Linear Regression with Rebalanced KNN Upsampling")
    # lr_reb_model, lr_reb_test_mape = build_lr(X_reb, X_test, y_reb, y_test)
    # print("")

    # print("Building No Bootstrap Model with no Upsampling")
    # no_bootstrap_model, no_bootstrap_test_mape = build_no_bootstrap(X_train, X_test, y_flat, y_test)
    # print("")
    
    # print("Building No Bootstrap Model with Simple Duplication")
    # no_bootstrap_sim_model, no_bootstrap_sim_test_mape = build_no_bootstrap(X_sim, X_test, y_sim, y_test, cv=5)
    # print("")

    # print("Building No Bootstrap Model with KNN Upsampling")
    # no_bootstrap_knn_model, no_bootstrap_knn_test_mape = build_no_bootstrap(X_knn, X_test, y_knn, y_test, cv=5)
    # print("")

    # print("Building No Bootstrap Model with Rebalanced KNN Upsampling")
    # no_bootstrap_reb_model, no_bootstrap_reb_test_mape = build_no_bootstrap(X_reb, X_test, y_reb, y_test, cv=5)
    # print("")

    # print("Building Bootstrap Model with no Upsampling")
    # bootstrap_model, bootstrap_test_mape = build_bootstrap(X_train, X_test, y_flat, y_test)
    # print("")
    
    # print("Building Bootstrap Model with Simple Duplication")
    # bootstrap_sim_model, bootstrap_sim_test_mape = build_bootstrap(X_sim, X_test, y_sim, y_test, cv=5)
    # print("")

    # print("Building Bootstrap Model with KNN Upsampling")
    # bootstrap_knn_model, bootstrap_knn_test_mape = build_bootstrap(X_knn, X_test, y_knn, y_test, cv=5)
    # print("")

    # print("Building Bootstrap Model with Rebalanced KNN Upsampling")
    # bootstrap_reb_model, bootstrap_reb_test_mape = build_bootstrap(X_reb, X_test, y_reb, y_test, cv=5)
    # print("")

    # print("Building XGBoost Model with no Upsampling")
    # xgboost_model, xgboost_test_mape = build_xgboost(X_train, X_test, y_flat, y_test)
    # print("")
    
    # print("Building XGBoost Model with Simple Duplication")
    # xgboost_sim_model, xgboost_sim_test_mape = build_xgboost(X_sim, X_test, y_sim, y_test, cv=5)
    # print("")

    # print("Building XGBoost Model with KNN Upsampling")
    # xgboost_knn_model, xgboost_knn_test_mape = build_xgboost(X_knn, X_test, y_knn, y_test, cv=5)
    # print("")
    
    # print("Building XGBoost Model with Rebalanced KNN Upsampling")
    # xgboost_reb_model, xgboost_reb_test_mape = build_xgboost(X_reb, X_test, y_reb, y_test, cv=5)
    # print("")

    # print("Building AdaBoost Model with no Upsampling")
    # adaboost_model, adaboost_test_mape = build_adaboost(X_train, X_test, y_flat, y_test)
    # print("")

    # print("Building AdaBoost Model with Simple Dulpication")
    # adaboost_sim_model, adaboost_sim_test_mape = build_adaboost(X_sim, X_test, y_sim, y_test)
    # print("")

    # print("Building AdaBoost Model with KNN Upsampling")
    # adaboost_knn_model, adaboost_knn_test_mape = build_adaboost(X_knn, X_test, y_knn, y_test)
    # print("")

    # print("Building AdaBoost Model with Rebalanced KNN Upsampling")
    # adaboost_reb_model, adaboost_reb_test_mape = build_adaboost(X_reb, X_test, y_reb, y_test)
    # print("")

    print("Building XGBoost with Bootstrapping and no Oversampling")
    xgboost_bootstrap_model, xgboost_bootstrap_mape = build_xgboost_with_bootstrap(X_train, X_test, y_flat, y_test)
    print("")

    print("Building XGBoost with Bootstrapping and Simple Duplication")
    xgboost_bootstrap_sim_model, xgboost_bootstrap_sim_mape = build_xgboost_with_bootstrap(X_sim, X_test, y_sim, y_test)
    print("")

    print("Building XGBoost with Bootstrapping and KNN Oversampling")
    xgboost_bootstrap_knn_model, xgboost_bootstrap_knn_mape = build_xgboost_with_bootstrap(X_knn, X_test, y_knn, y_test)
    print("")

    print("Building XGBoost with Bootstrapping and SMO")
    xgboost_bootstrap_reb_model, xgboost_bootstrap_reb_mape = build_xgboost_with_bootstrap(X_reb, X_test, y_reb, y_test)
    print("")

    print("Building AdaBoost with Bootstrapping and no Oversampling")
    adaboost_bootstrap_model, adaboost_bootstrap_mape = build_adaboost_with_bootstrap(X_train, X_test, y_flat, y_test)
    print("")

    print("Building AdaBoost with Bootstrapping and Simple Duplication")
    adaboost_bootstrap_sim_model, adaboost_bootstrap_sim_mape = build_adaboost_with_bootstrap(X_sim, X_test, y_sim, y_test)
    print("")

    print("Building AdaBoost with Bootstrapping and KNN Oversampling")
    adaboost_bootstrap_knn_model, adaboost_bootstrap_knn_mape = build_adaboost_with_bootstrap(X_knn, X_test, y_knn, y_test)
    print("")

    print("Building AdaBoost with Bootstrapping and SMO")
    adaboost_bootstrap_reb_model, adaboost_bootstrap_reb_mape = build_adaboost_with_bootstrap(X_reb, X_test, y_reb, y_test)
    print("")

    # # Row is model type
    # summary = pd.DataFrame([[lr_test_mape, lr_sim_test_mape, lr_knn_test_mape, lr_reb_test_mape],
    #                        [no_bootstrap_test_mape, no_bootstrap_sim_test_mape, no_bootstrap_knn_test_mape, no_bootstrap_reb_test_mape],
    #                        [bootstrap_test_mape, bootstrap_sim_test_mape, bootstrap_knn_test_mape, bootstrap_reb_test_mape],
    #                        [xgboost_test_mape, xgboost_sim_test_mape, xgboost_knn_test_mape, xgboost_reb_test_mape],
    #                        [adaboost_test_mape, adaboost_sim_test_mape, adaboost_knn_test_mape, adaboost_reb_test_mape],
    #                        [stacking_test_mape, stacking_sim_test_mape, stacking_knn_test_mape, stacking_reb_test_mape]],
    #                       index = ['LinearRegression', 'RF No Bootstrap', 'RF Bootstrap', 'XGBoost', 'AdaBoost', 'Stacking'],
    #                       columns = ['No Upsampling', 'Simple Duplication', 'KNN', 'Rebalanced KNN'])
    # print("MAPE Summary Table")
    # print(summary)
    # print("")

    # if args.model == 0:
    #     summary.to_excel('./Datasets/model_mape_qaly.xlsx', index=False, engine='openpyxl')
    #     print("Results Saved")
    #     models = [no_bootstrap_reb_model, bootstrap_reb_model, xgboost_reb_model, stacking_reb_model]
    # else:
    #     summary.to_excel('./Datasets/model_mape_hsc.xlsx', index=False, engine='openpyxl')
    #     print("Results Saved")
    #     models = [no_bootstrap_reb_model, bootstrap_reb_model, xgboost_reb_model, stacking_reb_model]
        

    # Generate all combinations (1 to 5 models)
    # all_combinations = []
    # for r in range(1, 5):  # From 1 model to 5 models
    #     combinations = list(itertools.combinations(models, r))
    #     all_combinations.extend(combinations)

    # # Convert each tuple into a list
    # all_combinations = [list(combo) for combo in all_combinations]

    # best_mape = 10
    # best_ensemble = []
    # for i in range(len(all_combinations)):
    #     current_mape = ensemble(all_combinations[i], X_test, y_test)
    #     if current_mape < best_mape:
    #         best_mape = current_mape
    #         best_ensemble = all_combinations[i]
    # print("Best Ensemble", best_ensemble)
    # print("Ensemble MAPE", best_mape)
        


    
    if args.variance:
        new_row = [no_bootstrap_reb_test_mape, bootstrap_reb_test_mape, xgboost_reb_test_mape]
        if args.model == 0:
            append_to_excel("./Datasets/variance_qaly.xlsx", new_row)
        else:
            append_to_excel("./Datasets/variance_hsc.xlsx", new_row)
            
        
        
    
    
    
    
        
    
    